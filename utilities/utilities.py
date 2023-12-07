import random
import string
from datetime import datetime

import openpyxl


def generate_email_with_timestamp():
    now = datetime.now()
    ts_str = now.strftime("""%Y_%m_%d_%H_%M_%S""")
    return f"{ts_str}@gmail.com"


def generate_random_string():
    char_count = 10
    random_string = "".join(
        random.choices(string.ascii_letters + string.digits, k=char_count)
    )
    return random_string


def get_cell_value(file_path, sheet_name, target_row_value, target_column_header):
    """

    :param file_path: Path to file containing data
    :param sheet_name: Sheetname (Sheet1, Sheet2, ...)
    :param target_row_value: Row value to look for
    :param target_column_header: Column value to look for
    :return: Based on given target_row_value and target_column_header, return its value
    """
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the specified sheet
    sheet = workbook[sheet_name]

    # Find the target column index based on the column header
    column_index = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == target_column_header:
            column_index = col
            break

    if column_index is None:
        print(f"Column '{target_column_header}' not found in the sheet.")
        return None

    # Find the target row index based on the row value
    row_index = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == target_row_value:
            row_index = row
            break

    if row_index is None:
        print(f"Row '{target_row_value}' not found in the sheet.")
        return None

    # Get the value in the specified cell
    cell_value = sheet.cell(row=row_index, column=column_index).value

    # Close the workbook
    workbook.close()

    return cell_value


def get_row_values(file_path, sheet_name, target_row_value):
    """
    Retrieve all cell values based on given target_row_value
    :param file_path:
    :param sheet_name:
    :param target_row_value:
    :return: Returns list of values based on given target_row_value
    """
    # To add final values retrieved
    row_values = []

    # Load Excel file and assign a sheet to read from
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    row_index = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == target_row_value:
            row_index = row
            break
    if row_index is None:
        print(f"{target_row_value} not found in file {file_path} ({sheet_name}).")

    for items in range(2, sheet.max_column + 1):
        row_values.append(sheet.cell(row=row_index, column=items).value)

    return row_values
